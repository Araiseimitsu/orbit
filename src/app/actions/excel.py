"""
Excel actions (local .xlsx/.xlsm files)
"""

from __future__ import annotations

import json
import logging
from datetime import datetime, date
from pathlib import Path
from typing import Any, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from ..core.registry import register_action

logger = logging.getLogger(__name__)


def _resolve_path(path_str: str, base_dir: Path) -> Path:
    if not path_str:
        raise ValueError("path は必須です")
    path = Path(path_str)
    if not path.is_absolute():
        path = base_dir / path
    return path


def _normalize_cell_value(value: Any) -> Any:
    """
    セルの値を正規化（datetime/date オブジェクトを文字列に変換）

    Args:
        value: セルの値（datetime, date, その他）

    Returns:
        正規化された値（datetime/date は YYYY/MM/DD 形式の文字列に変換）
    """
    if value is None:
        return ""

    if isinstance(value, datetime):
        # datetime オブジェクトを YYYY/MM/DD 形式の文字列に変換
        return value.strftime("%Y/%m/%d")

    if isinstance(value, date):
        # date オブジェクトを YYYY/MM/DD 形式の文字列に変換
        return value.strftime("%Y/%m/%d")

    # その他の型はそのまま返す
    return value


def _normalize_values(values: Any) -> list[list[Any]]:
    if values is None:
        raise ValueError("values は必須です")
    if isinstance(values, str):
        values_str = values.strip()
        if not values_str:
            raise ValueError("values は必須です")
        try:
            values = json.loads(values_str)
        except json.JSONDecodeError as e:
            raise ValueError("values は2次元配列(JSON)で指定してください") from e

    if not isinstance(values, list):
        raise ValueError("values は2次元配列で指定してください")
    if values and not all(isinstance(row, list) for row in values):
        raise ValueError("values は2次元配列で指定してください")
    return values


def _parse_values_with_header(
    values: list[list[Any]], header_row: bool
) -> dict[str, Any]:
    if not values:
        return {"headers": [], "rows": [], "raw": [], "row_count": 0, "col_count": 0}

    if header_row and len(values) >= 1:
        # ヘッダー行を正規化（datetime オブジェクトの可能性があるため）
        normalized_headers = [_normalize_cell_value(h) for h in values[0]]
        headers = [
            str(h) if h else f"col_{i}" for i, h in enumerate(normalized_headers)
        ]
        data_rows = values[1:]

        rows = []
        for row in data_rows:
            row_dict = {}
            for i, header in enumerate(headers):
                cell_value = row[i] if i < len(row) else ""
                # セル値を正規化（datetime/date を文字列に変換）
                normalized_value = _normalize_cell_value(cell_value)
                row_dict[header] = normalized_value
            rows.append(row_dict)

        # raw データも正規化
        normalized_raw = []
        for row in values:
            normalized_raw.append([_normalize_cell_value(cell) for cell in row])

        return {
            "headers": headers,
            "rows": rows,
            "raw": normalized_raw,
            "row_count": len(data_rows),
            "col_count": len(headers),
        }

    # ヘッダーなしの場合も正規化
    col_count = max(len(row) for row in values) if values else 0
    normalized_raw = []
    for row in values:
        normalized_raw.append([_normalize_cell_value(cell) for cell in row])

    return {
        "headers": [],
        "rows": [],
        "raw": normalized_raw,
        "row_count": len(values),
        "col_count": col_count,
    }


def _split_sheet_and_range(
    range_str: str | None, sheet: str | None
) -> tuple[str | None, str | None]:
    if not range_str:
        return sheet, None
    if "!" in range_str:
        sheet_from_range, range_part = range_str.split("!", 1)
        if not sheet:
            sheet = sheet_from_range
        return sheet, range_part
    return sheet, range_str


def _is_xlsm(path: Path) -> bool:
    return path.suffix.lower() == ".xlsm"


def _load_workbook_for_read(path: Path, data_only: bool) -> Workbook:
    if not path.exists():
        raise FileNotFoundError(f"Excel ファイルが見つかりません: {path}")
    return load_workbook(filename=path, data_only=data_only, keep_vba=_is_xlsm(path))


def _load_workbook_for_write(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(filename=path, keep_vba=_is_xlsm(path))
    path.parent.mkdir(parents=True, exist_ok=True)
    return Workbook()


def _get_sheet(wb: Workbook, sheet_name: str | None, create: bool) -> Any:
    if sheet_name:
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        if create:
            return wb.create_sheet(sheet_name)
        raise ValueError(f"シートが見つかりません: {sheet_name}")
    return wb.active


def _range_to_bounds(range_str: str) -> tuple[int, int, int, int]:
    try:
        return range_boundaries(range_str)
    except ValueError as e:
        raise ValueError(f"range の形式が不正です: {range_str}") from e


def _next_row_for_append(sheet) -> int:
    if sheet.max_row == 1 and all(cell.value is None for cell in sheet[1]):
        return 1
    return sheet.max_row + 1


def _calc_updated_range(
    start_row: int, start_col: int, row_count: int, col_count: int
) -> str:
    if row_count <= 0 or col_count <= 0:
        return ""
    end_row = start_row + row_count - 1
    end_col = start_col + col_count - 1
    start_cell = f"{get_column_letter(start_col)}{start_row}"
    end_cell = f"{get_column_letter(end_col)}{end_row}"
    return f"{start_cell}:{end_cell}"


@register_action(
    "excel_read",
    metadata={
        "title": "Excel 読み込み",
        "description": "Excelファイルからデータを読み取ります。",
        "category": "Excel",
        "color": "#22c55e",
        "params": [
            {
                "key": "path",
                "description": "Excelファイルパス",
                "required": True,
                "example": "data.xlsx",
            },
            {
                "key": "sheet",
                "description": "シート名（省略時はアクティブシート）",
                "required": False,
                "example": "Sheet1",
            },
            {
                "key": "range",
                "description": "取得範囲（例: 'A1:D10' または 'Sheet1!A1:D10'）",
                "required": True,
                "example": "A1:D10",
            },
            {
                "key": "header_row",
                "description": "1行目をヘッダーとして扱う",
                "required": False,
                "default": True,
                "example": "true",
            },
            {
                "key": "data_only",
                "description": "数式の結果を返す",
                "required": False,
                "default": True,
                "example": "true",
            },
        ],
        "outputs": [
            {"key": "headers", "description": "ヘッダー行"},
            {"key": "rows", "description": "データ行"},
            {"key": "raw", "description": "生データ"},
            {"key": "row_count", "description": "行数"},
            {"key": "col_count", "description": "列数"},
        ],
    },
)
async def action_excel_read(
    params: dict[str, Any], context: dict[str, Any]
) -> dict[str, Any]:
    """
    Excel からデータを読み取る

    params:
        path: Excel ファイルパス
        sheet: シート名（省略時はアクティブシート）
        range: 取得範囲（例: "A1:D10" または "Sheet1!A1:D10"）
        header_row: 1行目をヘッダーとして扱う (デフォルト: true)
        data_only: 数式の結果を返す (デフォルト: true)
    """
    path_str = params.get("path", "")
    sheet = params.get("sheet")
    range_str = params.get("range")
    header_row = params.get("header_row", True)
    data_only = params.get("data_only", True)

    if not range_str:
        raise ValueError("range は必須です (例: 'A1:D10')")

    base_dir = context.get("base_dir", Path.cwd())
    path = _resolve_path(path_str, base_dir)
    sheet, range_part = _split_sheet_and_range(range_str, sheet)

    wb = _load_workbook_for_read(path, data_only=bool(data_only))
    ws = _get_sheet(wb, sheet, create=False)

    min_col, min_row, max_col, max_row = _range_to_bounds(range_part)
    values = []
    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    ):
        values.append(list(row))

    parsed = _parse_values_with_header(values, bool(header_row))
    return {
        **parsed,
        "path": str(path),
        "sheet": ws.title,
        "range": range_part,
    }


@register_action(
    "excel_list_sheets",
    metadata={
        "title": "Excel シート一覧",
        "description": "Excelファイルのシート一覧を取得します。",
        "category": "Excel",
        "color": "#22c55e",
        "params": [
            {
                "key": "path",
                "description": "Excelファイルパス",
                "required": True,
                "example": "data.xlsx",
            }
        ],
        "outputs": [
            {"key": "sheets", "description": "シート名のリスト"},
            {"key": "path", "description": "ファイルパス"},
        ],
    },
)
async def action_excel_list_sheets(
    params: dict[str, Any], context: dict[str, Any]
) -> dict[str, Any]:
    """
    Excel のシート一覧を取得
    """
    path_str = params.get("path", "")
    base_dir = context.get("base_dir", Path.cwd())
    path = _resolve_path(path_str, base_dir)

    wb = _load_workbook_for_read(path, data_only=True)
    sheets = [{"title": name, "index": idx} for idx, name in enumerate(wb.sheetnames)]
    return {"path": str(path), "sheets": sheets}


@register_action(
    "excel_write",
    metadata={
        "title": "Excel 書き込み",
        "description": "Excelファイルの指定範囲にデータを書き込みます（上書き）。",
        "category": "Excel",
        "color": "#22c55e",
        "params": [
            {
                "key": "path",
                "description": "Excelファイルパス",
                "required": True,
                "example": "data.xlsx",
            },
            {
                "key": "sheet",
                "description": "シート名（省略時はアクティブシート）",
                "required": False,
                "example": "Sheet1",
            },
            {
                "key": "range",
                "description": "書き込み範囲。開始セルのみ（例: 'A1'）指定でデータサイズに合わせて自動拡張、範囲指定（例: 'A1:C2'）で指定範囲内のみ書き込み",
                "required": True,
                "example": "A1",
            },
            {
                "key": "values",
                "description": "2次元配列 or JSON文字列（例: {{ step_id.raw }} / AIの出力 text は {{ ai_generate_1.text | fromjson }}）",
                "required": True,
                "example": "{{ ai_generate_1.text | fromjson }}",
            },
        ],
        "outputs": [
            {"key": "path", "description": "ファイルパス"},
            {"key": "sheet", "description": "シート名"},
            {"key": "range", "description": "範囲"},
            {"key": "updated_range", "description": "更新範囲"},
            {"key": "updated_rows", "description": "更新行数"},
            {"key": "updated_columns", "description": "更新列数"},
        ],
    },
)
async def action_excel_write(
    params: dict[str, Any], context: dict[str, Any]
) -> dict[str, Any]:
    """
    Excel の指定範囲に書き込み（上書き）

    params:
        path: Excel ファイルパス
        sheet: シート名（省略時はアクティブシート）
        range: 書き込み範囲（例: "A1:C2" または "Sheet1!A1:C2"）
        values: 2次元配列 or JSON文字列
    """
    path_str = params.get("path", "")
    sheet = params.get("sheet")
    range_str = params.get("range")
    values = _normalize_values(params.get("values"))

    if not range_str:
        raise ValueError("range は必須です (例: 'A1:C2')")

    base_dir = context.get("base_dir", Path.cwd())
    path = _resolve_path(path_str, base_dir)
    sheet, range_part = _split_sheet_and_range(range_str, sheet)

    wb = _load_workbook_for_write(path)
    ws = _get_sheet(wb, sheet, create=True)

    min_col, min_row, max_col, max_row = _range_to_bounds(range_part)
    row_count = len(values)
    col_count = max((len(row) for row in values), default=0)

    if row_count == 0 or col_count == 0:
        raise ValueError("values は空でない2次元配列が必要です")

    # 開始セルのみ指定の場合（例: G1）、データサイズに合わせて書き込み
    # 範囲指定の場合（例: A1:C2）、指定範囲内のみ書き込み
    is_single_cell = min_col == max_col and min_row == max_row
    if not is_single_cell:
        range_rows = max_row - min_row + 1
        range_cols = max_col - min_col + 1
        if row_count > range_rows or col_count > range_cols:
            raise ValueError("values のサイズが range を超えています")

    for r_idx, row in enumerate(values):
        for c_idx, value in enumerate(row):
            ws.cell(row=min_row + r_idx, column=min_col + c_idx, value=value)

    wb.save(path)

    updated_range = _calc_updated_range(min_row, min_col, row_count, col_count)
    return {
        "path": str(path),
        "sheet": ws.title,
        "range": range_part,
        "updated_range": updated_range,
        "updated_rows": row_count,
        "updated_columns": col_count,
        "updated_cells": row_count * col_count,
    }


@register_action(
    "excel_append",
    metadata={
        "title": "Excel 追記",
        "description": "Excelファイルの最後に行を追加（追記）します。",
        "category": "Excel",
        "color": "#22c55e",
        "params": [
            {
                "key": "path",
                "description": "Excelファイルパス",
                "required": True,
                "example": "data.xlsx",
            },
            {
                "key": "sheet",
                "description": "シート名（省略時はアクティブシート）",
                "required": False,
                "example": "Sheet1",
            },
            {
                "key": "values",
                "description": "2次元配列 or JSON文字列（AIの出力 text を使う場合は {{ ai_generate_1.text | fromjson }} のように変換）",
                "required": True,
                "example": "{{ ai_generate_1.text | fromjson }}",
            },
            {
                "key": "start_cell",
                "description": "追記開始列を指定（例: 'B1'。行番号は無視）",
                "required": False,
                "default": "A1",
                "example": "A1",
            },
        ],
        "outputs": [
            {"key": "path", "description": "ファイルパス"},
            {"key": "sheet", "description": "シート名"},
            {"key": "appended_range", "description": "追記範囲"},
            {"key": "appended_rows", "description": "追記行数"},
            {"key": "appended_columns", "description": "追記列数"},
        ],
    },
)
async def action_excel_append(
    params: dict[str, Any], context: dict[str, Any]
) -> dict[str, Any]:
    """
    Excel に行を追加（追記）

    params:
        path: Excel ファイルパス
        sheet: シート名（省略時はアクティブシート）
        values: 2次元配列 or JSON文字列
        start_cell: 追記開始列を指定（例: "B1"。行番号は無視）
    """
    path_str = params.get("path", "")
    sheet = params.get("sheet")
    values = _normalize_values(params.get("values"))
    start_cell = params.get("start_cell", "A1")

    base_dir = context.get("base_dir", Path.cwd())
    path = _resolve_path(path_str, base_dir)

    wb = _load_workbook_for_write(path)
    ws = _get_sheet(wb, sheet, create=True)

    try:
        start_col, _, _, _ = range_boundaries(start_cell)
    except ValueError as e:
        raise ValueError(f"start_cell の形式が不正です: {start_cell}") from e

    row_count = len(values)
    col_count = max((len(row) for row in values), default=0)
    if row_count == 0 or col_count == 0:
        raise ValueError("values は空でない2次元配列が必要です")

    start_row = _next_row_for_append(ws)
    for r_idx, row in enumerate(values):
        for c_idx, value in enumerate(row):
            ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)

    wb.save(path)

    updated_range = _calc_updated_range(start_row, start_col, row_count, col_count)
    return {
        "path": str(path),
        "sheet": ws.title,
        "start_row": start_row,
        "updated_range": updated_range,
        "appended_rows": row_count,
        "appended_columns": col_count,
        "appended_cells": row_count * col_count,
    }
